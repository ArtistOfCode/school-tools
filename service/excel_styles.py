from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Color
from pptx.dml.color import RGBColor
from pptx.enum.text import MSO_VERTICAL_ANCHOR, PP_ALIGN

__bold = Font(bold=True)
__center = Alignment(horizontal='center')
__fill = PatternFill('solid', fgColor=Color('e3e3e3'))
__thin_side = Side(style='thin', color='000000')
__thin_border = Border(__thin_side, __thin_side, __thin_side, __thin_side)


def set_title_cell(cell: Cell, value) -> Cell:
    return set_cell(cell, value, __bold, __center, __fill)


def set_float_cell(cell: Cell, value) -> Cell:
    cell.number_format = '0.00'
    return set_cell(cell, value)


def set_cell(cell: Cell, value, font: Font = None, align: Alignment = None, fill: PatternFill = None,
             border: Border = __thin_border) -> Cell:
    cell.value = value
    if font is not None: cell.font = font
    if align is not None: cell.alignment = align
    if fill is not None: cell.fill = fill
    if border is not None: cell.border = border
    return cell


def set_center_cell(cell, value: str, color=None):
    cell.text = value
    cell.vertical_anchor = MSO_VERTICAL_ANCHOR.MIDDLE
    cell.text_frame.paragraphs[0].alignment = PP_ALIGN.CENTER
    if color is not None: cell.text_frame.paragraphs[0].font.color.rgb = RGBColor.from_string(color)


class CellIndex:

    def __init__(self, row=1):
        self.value = row

    def next(self, step=1):
        self.value += step
